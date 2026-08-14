"""XLSX 连接器（Step 18）：openpyxl 读取 → pyarrow 表注册。

公式注入扫描价值最高（电子表格场景）；混合类型列自动推断失败时回退全字符串
（ADR-019）。read_batches 内存切片（MVP 预算 1e6 行内，ADR-019）。
"""

from __future__ import annotations

from collections.abc import Iterator

import pyarrow as pa
from openpyxl import load_workbook

from datasentry_core.connectors.base import FrameBatch
from datasentry_core.connectors.errors import ConnectorError
from datasentry_core.connectors.file_based import FileDataHandle
from datasentry_core.connectors.spec import DataSourceSpec, DataSourceType

#: 整 sheet 入内存的行预算（ADR-019 MVP 语义，Step 73/ADR-073 显式化）
_XLSX_ROW_BUDGET = 1_000_000


def _rows_to_table(names: list[str], rows: list[list[object]]) -> pa.Table:
    """行值列表 → pyarrow 表；混合类型列自动推断失败时回退全字符串（ADR-019）。"""
    pylist = [dict(zip(names, r, strict=False)) for r in rows]
    try:
        return pa.Table.from_pylist(pylist)
    except (pa.ArrowInvalid, pa.ArrowTypeError):
        stringified = [
            {name: (str(value) if value is not None else None) for name, value in r.items()}
            for r in pylist
        ]
        return pa.Table.from_pylist(
            stringified,
            schema=pa.schema([(name, pa.string()) for name in names]),
        )


class XlsxDataHandle(FileDataHandle):
    """XLSX 数据句柄。"""

    def _load_rows(self) -> tuple[list[str], list[list[object]]]:
        """读取指定 sheet 为 (列名, 行值)（header_row 可配，默认第 0 行）。

        data_only=False：公式单元格返回公式文本（公式注入可检测，ADR-019）；
        计算结果缓存读取归 V1。
        """
        sheet_ref = self._spec.options.get("sheet")
        header_row = int(self._spec.options.get("header_row", 0))
        workbook = load_workbook(self._path, read_only=True, data_only=False)
        try:
            if sheet_ref is None:
                sheet = workbook.worksheets[0]
            elif isinstance(sheet_ref, int):
                sheet = workbook.worksheets[sheet_ref]
            else:
                sheet = workbook[str(sheet_ref)]
            rows: list[tuple[object, ...]] = []
            for row in sheet.iter_rows(values_only=True):
                if len(rows) >= _XLSX_ROW_BUDGET:
                    raise ConnectorError(
                        f"xlsx sheet exceeds row budget {_XLSX_ROW_BUDGET} "
                        "(ADR-019/ADR-073); split the sheet or use --sampling"
                    )
                rows.append(row)
        finally:
            workbook.close()
        if header_row >= len(rows):
            return [], []
        names = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[header_row])]
        data = [list(r) for r in rows[header_row + 1 :]]
        return names, data

    def _ensure_view(self) -> None:
        if self._view_ready:
            return
        names, rows = self._load_rows()
        if not names:
            raise ValueError("xlsx sheet is empty")
        if not rows:
            table = pa.table({name: pa.array([], type=pa.null()) for name in names})
        else:
            table = _rows_to_table(names, rows)
        self._executor.register("data", table)
        self._view_ready = True

    def read_batches(self, batch_size: int = 65536) -> Iterator[FrameBatch]:
        self._check_open()
        if batch_size < 1:
            raise ValueError("batch_size must be >= 1")
        names, rows = self._load_rows()
        if not names:
            return
        for idx in range(0, len(rows), batch_size):
            table = _rows_to_table(names, rows[idx : idx + batch_size])
            yield FrameBatch(
                source=self._spec,
                table=table,
                batch_index=idx // batch_size,
                row_offset=idx,
            )


class XlsxConnector:
    """XLSX 连接器（7.1）。"""

    connector_id = "xlsx"
    display_name = "XLSX"

    def supports(self, source: DataSourceSpec) -> bool:
        return source.source_type == DataSourceType.XLSX and source.path is not None

    def open(self, source: DataSourceSpec) -> XlsxDataHandle:
        return XlsxDataHandle(source)

    def close(self) -> None:
        pass


__all__ = ["XlsxConnector", "XlsxDataHandle"]
