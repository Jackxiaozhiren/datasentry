"""修复引擎（12.5/15 章 MVP 子集，Step 19，ADR-020）。

闭环：propose（Issue → 提案）→ preview（统计面板 + 检测器重跑）→
apply（修复副本 + before artifact）→ rollback（artifact 全量重建）。

设计约束（ADR-020）：
- 只支持确定性、值级操作（trim/case/token→null/set_null/clip）；
  impute 等推断类归 V1（伪造数据风险）
- 原文件永不变；产物在 <workspace>/.datasentry/repairs/
- 回滚 = before artifact 全量重建（operation log 仅样本）
- rule_failures 前后对比 = 重跑同检测器（副本上）
"""

from __future__ import annotations

import json
import shutil
import tempfile
import uuid
from pathlib import Path
from typing import Any

import pyarrow as pa
import pyarrow.csv as pa_csv
import pyarrow.parquet as pq

from datasentry_core.connectors import (
    DataSourceSpec,
    DataSourceType,
    default_registry,
)
from datasentry_core.detectors import DetectionContext, DetectorRegistry
from datasentry_core.detectors.common import quote_ident, quote_literal, quote_re
from datasentry_core.models.enums import (
    RepairOperation,
    RepairProposalStatus,
    RepairRunStatus,
    RiskLevel,
)
from datasentry_core.models.issue import Issue
from datasentry_core.models.repair import (
    RepairOperationRecord,
    RepairPreview,
    RepairProposal,
    RepairRun,
    RowBeforeAfter,
)
from datasentry_core.storage.paths import project_repairs_dir

# 缺失标记（与 suspicious_missing_token 检测器一致）
_MISSING_TOKENS = (
    "na",
    "n/a",
    "null",
    "none",
    "-",
    "?",
    "unknown",
    "missing",
    "todo",
    "tbd",
    "n.a.",
)
_ISO_DT_RE = r"^\d{4}-\d{2}-\d{2}$"
_OPERATION_LOG_CAP = 500
_EXAMPLE_CAP = 10

# Issue 类型 → 修复操作（MVP 确定性值级子集；其余检测器不自动提案）
_PROPOSAL_MAP: dict[str, RepairOperation] = {
    "leading_or_trailing_whitespace": RepairOperation.TRIM_WHITESPACE,
    "inconsistent_case": RepairOperation.NORMALIZE_CASE,
    "suspicious_missing_token": RepairOperation.REPLACE_MISSING_TOKEN,
    "invalid_date": RepairOperation.SET_NULL,
    "impossible_date": RepairOperation.SET_NULL,
}
# CLIP_VALUE：仅当 evidence 提供 lower/upper 边界（数值离群类）
_CLIP_ISSUE_TYPES = frozenset({"iqr_outlier", "percentile_outlier", "modified_zscore"})

_RATIONALE: dict[RepairOperation, str] = {
    RepairOperation.TRIM_WHITESPACE: "strip leading/trailing whitespace",
    RepairOperation.NORMALIZE_CASE: "lowercase values to a single canonical form",
    RepairOperation.REPLACE_MISSING_TOKEN: "replace missing stand-ins with NULL",
    RepairOperation.SET_NULL: "set invalid values to NULL (missing semantics)",
    RepairOperation.CLIP_VALUE: "clip values to the detected outlier bounds",
}


def _after_expr(operation: RepairOperation, column: str, params: dict[str, Any]) -> str:
    """修复操作的 SQL 表达式（只读视图上计算 after 值）。"""
    q = quote_ident(column)
    if operation == RepairOperation.TRIM_WHITESPACE:
        return f"trim({q})"
    if operation == RepairOperation.NORMALIZE_CASE:
        return f"lower({q})"
    if operation == RepairOperation.REPLACE_MISSING_TOKEN:
        tokens = ", ".join(quote_literal(t) for t in _MISSING_TOKENS)
        return f"CASE WHEN lower(trim({q})) IN ({tokens}) THEN NULL ELSE {q} END"
    if operation == RepairOperation.SET_NULL:
        return (
            f"CASE WHEN {q} IS NOT NULL "
            f"AND NOT regexp_matches(trim({q}), {quote_re(_ISO_DT_RE)}) "
            f"OR ({q} IS NOT NULL AND regexp_matches(trim({q}), {quote_re(_ISO_DT_RE)}) "
            f"AND try_strptime(trim({q}), '%Y-%m-%d') IS NULL) "
            f"THEN NULL ELSE {q} END"
        )
    if operation == RepairOperation.CLIP_VALUE:
        lower = float(params["lower"])
        upper = float(params["upper"])
        return f"CASE WHEN {q} < {lower} THEN {lower} WHEN {q} > {upper} THEN {upper} ELSE {q} END"
    raise ValueError(f"unsupported repair operation: {operation}")


class RepairEngine:
    """确定性修复引擎（15 章 MVP 子集）。"""

    def propose(self, issue: Issue, context: DetectionContext) -> RepairProposal | None:
        """Issue → 修复提案；不支持的 issue 返回 None。

        融合后 Issue.issue_type 是家族名（string_format 等），原始类型在
        detector_ids（与检测器 issue_type 同名），按可修复优先级挑选。
        """
        bounds = self._clip_bounds(issue)
        clip_ok = bounds is not None and len(issue.columns) == 1
        source: str | None = None
        operation: RepairOperation | None = None
        for detector_id in issue.detector_ids:
            if detector_id in _PROPOSAL_MAP:
                source = detector_id
                operation = _PROPOSAL_MAP[detector_id]
                break
            if clip_ok and detector_id in _CLIP_ISSUE_TYPES:
                source = detector_id
                operation = RepairOperation.CLIP_VALUE
                break
        if source is None or operation is None:
            return None
        columns = list(issue.columns)
        if not columns:
            return None
        params: dict[str, Any] = {}
        if operation == RepairOperation.CLIP_VALUE:
            assert bounds is not None
            params = {"lower": bounds[0], "upper": bounds[1]}
        affected = self._affected_rows(context, operation, columns, params)
        if affected <= 0:
            return None
        return RepairProposal(
            proposal_id=f"prop_{uuid.uuid4().hex[:12]}",
            issue_id=issue.id,
            issue_type=source,
            operation=operation,
            target_columns=columns,
            parameters=params,
            rationale=_RATIONALE[operation],
            evidence_ids=[e.evidence_id for e in issue.evidence],
            risk_level=self._risk_level(operation),
            reversibility="fully_reversible"
            if operation != RepairOperation.SET_NULL
            else "partially_reversible",
            estimated_rows_changed=affected,
            status=RepairProposalStatus.PROPOSED,
        )

    @staticmethod
    def _clip_bounds(issue: Issue) -> tuple[float, float] | None:
        for evidence in issue.evidence:
            lower = evidence.data.get("lower")
            upper = evidence.data.get("upper")
            if lower is not None and upper is not None:
                return float(lower), float(upper)
        return None

    @staticmethod
    def _risk_level(operation: RepairOperation) -> RiskLevel:
        if operation in (
            RepairOperation.SET_NULL,
            RepairOperation.CLIP_VALUE,
            RepairOperation.REPLACE_MISSING_TOKEN,
        ):
            return RiskLevel.MEDIUM
        return RiskLevel.LOW

    @staticmethod
    def _affected_rows(
        context: DetectionContext,
        operation: RepairOperation,
        columns: list[str],
        params: dict[str, object],
    ) -> int:
        """估算受影响行数（任一目标列变化即计）。"""
        total = 0
        for column in columns:
            q = quote_ident(column)
            expr = _after_expr(operation, column, params)
            table = context.handle.sql_aggregate(
                f"SELECT count(*) AS n FROM data WHERE {q} IS DISTINCT FROM ({expr})"
            ).table
            total += int(table.column("n").to_pylist()[0])
        return total

    def preview(
        self,
        proposal: RepairProposal,
        context: DetectionContext,
        registry: DetectorRegistry,
    ) -> RepairPreview:
        """预览：统计面板 + 检测器重跑前后对比（临时副本，无副作用）。"""
        before_counts = self._rule_failures(proposal, context, registry)
        stats = self._stats(context, proposal)
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp) / ("preview" + self._suffix(context))
            after_table = self._after_table(context, proposal)
            self._write_table(tmp_path, after_table, context)
            after_handle = default_registry().open(
                DataSourceSpec(
                    source_type=context.handle.source_type,
                    path=tmp_path,
                    options={"dataset_id": context.dataset_id},
                )
            )
            try:
                after_ctx = DetectionContext(
                    dataset_id=context.dataset_id,
                    table_name=None,
                    columns=after_handle.schema().column_names,
                    handle=after_handle,
                )
                after_counts = self._rule_failures(proposal, after_ctx, registry)
            finally:
                after_handle.close()
        row_count = max(context.handle.count_rows(), 1)
        return RepairPreview(
            proposal_id=proposal.proposal_id,
            rows_changed=proposal.estimated_rows_changed,
            rows_changed_ratio=round(min(1.0, proposal.estimated_rows_changed / row_count), 6),
            null_delta=stats["null_delta"],
            unique_delta=stats["unique_delta"],
            rule_failures_before=before_counts,
            rule_failures_after=after_counts,
            changed_examples=self._examples(context, proposal),
        )

    def apply(
        self,
        proposal: RepairProposal,
        context: DetectionContext,
        workspace: Path,
        source_scan_run_id: str | None = None,
    ) -> RepairRun:
        """应用：写修复副本（<run_id><ext>）+ before artifact（.before<ext>）。"""
        repairs_dir = project_repairs_dir(workspace)
        repairs_dir.mkdir(parents=True, exist_ok=True)
        run_id = f"rep_{uuid.uuid4().hex[:12]}"
        source_path = context.handle.source_path
        if source_path is None or isinstance(source_path, str) or not source_path.exists():
            raise FileNotFoundError("repair requires an on-disk source file")
        suffix = self._suffix(context)
        artifact_path = repairs_dir / f"{run_id}.before{suffix}"
        output_path = repairs_dir / f"{run_id}{suffix}"
        shutil.copy2(source_path, artifact_path)
        after_table = self._after_table(context, proposal)
        self._write_table(output_path, after_table, context)
        fingerprint_before = context.handle.fingerprint()
        after_handle = default_registry().open(
            DataSourceSpec(
                source_type=context.handle.source_type,
                path=output_path,
                options={"dataset_id": context.dataset_id},
            )
        )
        try:
            fingerprint_after = after_handle.fingerprint()
        finally:
            after_handle.close()
        operations = self._operation_records(context, proposal)
        return RepairRun(
            id=run_id,
            dataset_id=context.dataset_id,
            proposal_id=proposal.proposal_id,
            source_scan_run_id=source_scan_run_id,
            fingerprint_before=fingerprint_before.file_sha256 or "",
            fingerprint_after=fingerprint_after.file_sha256 or "",
            operations=operations,
            rollback_artifact=str(artifact_path),
            status=RepairRunStatus.APPLIED,
        )

    def rollback(self, run: RepairRun, workspace: Path) -> RepairRun:
        """回滚：before artifact 全量重建修复副本（<id>.rolled_back<ext>）。"""
        if run.rollback_artifact is None:
            raise ValueError("run has no rollback artifact")
        artifact = Path(run.rollback_artifact)
        if not artifact.exists():
            raise FileNotFoundError(f"rollback artifact not found: {artifact}")
        repairs_dir = project_repairs_dir(workspace)
        # artifact 命名 <run_id>.before<suffix>，解析副本后缀
        prefix = f"{run.id}.before"
        if artifact.name.startswith(prefix):
            suffix = artifact.name[len(prefix) :]
        else:
            suffix = artifact.suffix
        output_path = repairs_dir / f"{run.id}.rolled_back{suffix}"
        shutil.copy2(artifact, output_path)
        return run.model_copy(update={"status": RepairRunStatus.ROLLED_BACK})

    # ---- 内部 -----------------------------------------------------------

    def _after_table(self, context: DetectionContext, proposal: RepairProposal) -> pa.Table:
        """修复后的全量表（SQL 表达式，其余列原样）。"""
        select = []
        for column in context.columns:
            if column in proposal.target_columns:
                q = quote_ident(column)
                select.append(
                    f"{_after_expr(proposal.operation, column, proposal.parameters)} AS {q}"
                )
            else:
                select.append(quote_ident(column))
        return context.handle.sql_aggregate(f"SELECT {', '.join(select)} FROM data").table

    def _stats(
        self, context: DetectionContext, proposal: RepairProposal
    ) -> dict[str, dict[str, int]]:
        """null_delta / unique_delta（修复后 − 修复前）。"""
        null_delta: dict[str, int] = {}
        unique_delta: dict[str, int] = {}
        for column in proposal.target_columns:
            q = quote_ident(column)
            before = context.handle.sql_aggregate(
                f"SELECT sum({q} IS NULL) AS n, count(DISTINCT {q}) AS u FROM data"
            ).table
            expr = _after_expr(proposal.operation, column, proposal.parameters)
            after = context.handle.sql_aggregate(
                f"SELECT sum(({expr}) IS NULL) AS n, count(DISTINCT ({expr})) AS u FROM data"
            ).table
            null_delta[column] = int(after.column("n").to_pylist()[0]) - int(
                before.column("n").to_pylist()[0]
            )
            unique_delta[column] = int(after.column("u").to_pylist()[0]) - int(
                before.column("u").to_pylist()[0]
            )
        return {"null_delta": null_delta, "unique_delta": unique_delta}

    def _examples(
        self, context: DetectionContext, proposal: RepairProposal
    ) -> list[RowBeforeAfter]:
        """前 _EXAMPLE_CAP 个变化行（列级 before/after 示例）。"""
        examples: list[RowBeforeAfter] = []
        for column in proposal.target_columns:
            q = quote_ident(column)
            expr = _after_expr(proposal.operation, column, proposal.parameters)
            table = context.handle.sql_aggregate(
                f"SELECT {q} AS before_value, ({expr}) AS after_value FROM data "
                f"WHERE {q} IS DISTINCT FROM ({expr}) LIMIT {_EXAMPLE_CAP}"
            ).table
            before_values = table.column("before_value").to_pylist()
            after_values = table.column("after_value").to_pylist()
            for i, (before, after) in enumerate(zip(before_values, after_values, strict=True)):
                examples.append(
                    RowBeforeAfter(
                        row_id=f"{column}:{i}",
                        column=column,
                        before=before,
                        after=after,
                        reason=_RATIONALE[proposal.operation],
                    )
                )
        return examples

    def _rule_failures(
        self,
        proposal: RepairProposal,
        context: DetectionContext,
        registry: DetectorRegistry,
    ) -> dict[str, int]:
        """目标检测器在（原始/修复副本）上的候选数。"""
        detector_id = _DETECTOR_FOR_ISSUE.get(proposal.issue_type, "")
        if not detector_id:
            return {}
        try:
            detector = registry.get(detector_id)
        except KeyError:
            return {}
        if not detector.supports(context):
            return {}
        return {proposal.issue_type: len(detector.detect(context))}

    def _operation_records(
        self, context: DetectionContext, proposal: RepairProposal
    ) -> list[RepairOperationRecord]:
        """行级 before/after 记录样本（前 _OPERATION_LOG_CAP 条，回滚不依赖）。"""
        records: list[RepairOperationRecord] = []
        for column in proposal.target_columns:
            q = quote_ident(column)
            expr = _after_expr(proposal.operation, column, proposal.parameters)
            table = context.handle.sql_aggregate(
                f"SELECT {q} AS before_value, ({expr}) AS after_value FROM data "
                f"WHERE {q} IS DISTINCT FROM ({expr}) LIMIT {_OPERATION_LOG_CAP}"
            ).table
            before_values = table.column("before_value").to_pylist()
            after_values = table.column("after_value").to_pylist()
            for i, (before, after) in enumerate(zip(before_values, after_values, strict=True)):
                records.append(
                    RepairOperationRecord(
                        row_id=f"{column}:{i}",
                        column=column,
                        operation=proposal.operation,
                        before=before,
                        after=after,
                    )
                )
        return records

    @staticmethod
    def _suffix(context: DetectionContext) -> str:
        mapping = {
            DataSourceType.CSV: ".csv",
            DataSourceType.PARQUET: ".parquet",
            DataSourceType.JSONL: ".jsonl",
            DataSourceType.XLSX: ".xlsx",
        }
        return mapping.get(context.handle.source_type, ".csv")

    @staticmethod
    def _read_table(path: Path, source_type: DataSourceType) -> pa.Table:
        """修复工件读取（与 _write_table 格式对称）。"""
        if source_type == DataSourceType.PARQUET:
            return pq.read_table(path)
        if source_type == DataSourceType.JSONL:
            rows = [
                json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line
            ]
            return pa.Table.from_pylist(rows)
        if source_type == DataSourceType.XLSX:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True)
            ws = wb.active
            cols = [c.value for c in next(ws.iter_rows())]
            rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
            return pa.Table.from_pylist(
                [
                    dict(zip(cols, row, strict=False))
                    for row in rows
                    if row and any(v is not None for v in row)
                ]
            )
        return pa_csv.read_csv(path)

    @staticmethod
    def table_diff(
        before_path: Path, after_path: Path, source_type: DataSourceType
    ) -> tuple[list[str], list[list[object]], list[list[object]], list[int]]:
        """V43：修复工件 diff——(列名, before 行, after 行, 变更行索引)。"""
        before = RepairEngine._read_table(before_path, source_type)
        after = RepairEngine._read_table(after_path, source_type)
        columns = list(after.column_names)
        before_rows = [[row.get(name) for name in columns] for row in before.to_pylist()]
        after_rows = [[row.get(name) for name in columns] for row in after.to_pylist()]
        changed = [
            i for i, (b, a) in enumerate(zip(before_rows, after_rows, strict=False)) if b != a
        ]
        return columns, before_rows, after_rows, changed

    @staticmethod
    def _write_table(path: Path, table: pa.Table, context: DetectionContext) -> None:
        """修复副本写入（格式与源一致）。"""
        source_type = context.handle.source_type
        if source_type == DataSourceType.PARQUET:
            pq.write_table(table, path)
        elif source_type == DataSourceType.JSONL:
            with path.open("w", encoding="utf-8") as fh:
                for row in table.to_pylist():
                    fh.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")
        elif source_type == DataSourceType.XLSX:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.append(table.column_names)
            for row in table.to_pylist():
                ws.append([row.get(name) for name in table.column_names])
            wb.save(path)
        else:
            pa_csv.write_csv(table, path)


# Issue type → 检测器 id（rule_failures 重跑目标）
_DETECTOR_FOR_ISSUE: dict[str, str] = {
    "leading_or_trailing_whitespace": "leading_or_trailing_whitespace",
    "inconsistent_case": "inconsistent_case",
    "suspicious_missing_token": "suspicious_missing_token",
    "invalid_date": "invalid_date",
    "impossible_date": "impossible_date",
    "iqr_outlier": "iqr_outlier",
    "percentile_outlier": "percentile_outlier",
    "modified_zscore": "modified_zscore",
}
