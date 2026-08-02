"""Step 18 文件型连接器测试（Parquet/JSONL/XLSX，7.1 + ADR-019）。"""

from __future__ import annotations

import json
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq
import pytest
from openpyxl import Workbook

from datasentry_core.connectors import (
    CsvConnector,
    DataSourceSpec,
    DataSourceType,
    JsonlConnector,
    ParquetConnector,
    UnsupportedFormatError,
    XlsxConnector,
    default_registry,
)
from datasentry_core.connectors.errors import DataSourceNotFoundError
from datasentry_core.connectors.file_based import FileDataHandle
from datasentry_core.detectors import DetectionContext
from datasentry_core.detectors.initial.missing import ExcessiveNullRateDetector


def _write_parquet(path: Path) -> None:
    table = pa.table(
        {
            "name": ["alice", "bob", None, "dave", "eve"],
            "age": [30, 25, None, 41, 22],
        }
    )
    pq.write_table(table, path)


def _write_jsonl(path: Path) -> None:
    lines = [
        {"name": "alice", "age": 30},
        {"name": "bob", "age": 25},
        {"name": None, "age": None},
        {"name": "dave", "age": 41},
        {"name": "eve", "age": 22},
    ]
    path.write_text("\n".join(json.dumps(line) for line in lines) + "\n", encoding="utf-8")


def _write_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["name", "age"])
    ws.append(["alice", 30])
    ws.append(["bob", 25])
    ws.append([None, None])
    ws.append(["dave", 41])
    ws.append(["eve", 22])
    wb.save(path)


class TestParquetConnector:
    def test_open_and_query(self, tmp_path: Path) -> None:
        p = tmp_path / "t.parquet"
        _write_parquet(p)
        spec = DataSourceSpec(
            source_type=DataSourceType.PARQUET, path=p, options={"dataset_id": "t"}
        )
        handle = ParquetConnector().open(spec)
        try:
            assert handle.count_rows() == 5
            assert [c.name for c in handle.schema().columns] == ["name", "age"]
            sample = handle.read_sample(2)
            assert sample.table.num_rows == 2
            batch = next(handle.read_batches(batch_size=2))
            assert batch.table.num_rows == 2
            assert len(list(handle.read_batches(batch_size=2))) == 3
            fp = handle.fingerprint()
            assert fp.file_sha256 is not None
            assert fp.row_count == 5
            assert len(handle.warnings()) == 0
        finally:
            handle.close()

    def test_detectors_work(self, tmp_path: Path) -> None:
        p = tmp_path / "t.parquet"
        _write_parquet(p)
        spec = DataSourceSpec(
            source_type=DataSourceType.PARQUET, path=p, options={"dataset_id": "t"}
        )
        handle = ParquetConnector().open(spec)
        try:
            ctx = DetectionContext(
                dataset_id="t",
                table_name=None,
                columns=handle.schema().column_names,
                handle=handle,
            )
            candidates = ExcessiveNullRateDetector().detect(ctx)
            assert len(candidates) == 2
        finally:
            handle.close()


class TestJsonlConnector:
    def test_open_and_query(self, tmp_path: Path) -> None:
        p = tmp_path / "t.jsonl"
        _write_jsonl(p)
        spec = DataSourceSpec(source_type=DataSourceType.JSONL, path=p, options={"dataset_id": "t"})
        handle = JsonlConnector().open(spec)
        try:
            assert handle.count_rows() == 5
            assert [c.name for c in handle.schema().columns] == ["name", "age"]
            batches = list(handle.read_batches(batch_size=2))
            assert len(batches) == 3
            assert batches[0].row_offset == 0
            assert batches[2].row_offset == 4
        finally:
            handle.close()


class TestXlsxConnector:
    def test_open_and_query(self, tmp_path: Path) -> None:
        p = tmp_path / "t.xlsx"
        _write_xlsx(p)
        spec = DataSourceSpec(source_type=DataSourceType.XLSX, path=p, options={"dataset_id": "t"})
        handle = XlsxConnector().open(spec)
        try:
            assert handle.count_rows() == 5
            assert [c.name for c in handle.schema().columns] == ["name", "age"]
            types = {c.name: c.physical_type for c in handle.schema().columns}
            assert types["age"].startswith("INTEGER") or types["age"].startswith("BIGINT")
            fp = handle.fingerprint()
            assert fp.row_count == 5
        finally:
            handle.close()

    def test_formula_injection_warning(self, tmp_path: Path) -> None:
        p = tmp_path / "f.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["note"])
        ws.append(["=SUM(A1:A2)"])
        ws.append(["plain"])
        wb.save(p)
        spec = DataSourceSpec(source_type=DataSourceType.XLSX, path=p, options={"dataset_id": "f"})
        handle = XlsxConnector().open(spec)
        try:
            warnings = handle.warnings()
            assert len(warnings) == 1
            assert warnings[0].column == "note"
            assert warnings[0].row == 0
        finally:
            handle.close()

    def test_mixed_types_fallback_to_string(self, tmp_path: Path) -> None:
        p = tmp_path / "m.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["code"])
        ws.append([1])
        ws.append(["abc"])
        wb.save(p)
        spec = DataSourceSpec(source_type=DataSourceType.XLSX, path=p, options={"dataset_id": "m"})
        handle = XlsxConnector().open(spec)
        try:
            assert handle.count_rows() == 2
            assert handle.schema().columns[0].physical_type.upper() == "VARCHAR"
        finally:
            handle.close()


class TestFileHandleEdgeCases:
    """FileDataHandle 共享实现未覆盖分支（Step 22 覆盖率补齐）。"""

    def test_missing_path_raises(self, tmp_path: Path) -> None:
        spec = DataSourceSpec(source_type=DataSourceType.PARQUET, path=None, options={})
        with pytest.raises(DataSourceNotFoundError, match="requires a path"):
            ParquetConnector().open(spec)

    def test_read_sample_none_mode(self, tmp_path: Path) -> None:
        p = tmp_path / "t.parquet"
        _write_parquet(p)
        spec = DataSourceSpec(source_type=DataSourceType.PARQUET, path=p, options={})
        handle = ParquetConnector().open(spec)
        try:
            sample = handle.read_sample(2, method="none")
            assert sample.table.num_rows == 2
        finally:
            handle.close()

    def test_read_sample_time_based_without_column_falls_back(self, tmp_path: Path) -> None:
        p = tmp_path / "t.parquet"
        _write_parquet(p)
        spec = DataSourceSpec(source_type=DataSourceType.PARQUET, path=p, options={})
        handle = ParquetConnector().open(spec)
        try:
            sample = handle.read_sample(2, method="time_based")
            assert sample.table.num_rows == 2
        finally:
            handle.close()

    def test_read_sample_time_based_with_column(self, tmp_path: Path) -> None:
        p = tmp_path / "t.parquet"
        _write_parquet(p)
        spec = DataSourceSpec(
            source_type=DataSourceType.PARQUET,
            path=p,
            options={"time_column": "age"},
        )
        handle = ParquetConnector().open(spec)
        try:
            sample = handle.read_sample(2, method="time_based")
            assert sample.table.num_rows == 2
            assert sample.table.column("age").to_pylist() == [22, 25]
        finally:
            handle.close()

    def test_read_sample_invalid_n(self, tmp_path: Path) -> None:
        p = tmp_path / "t.parquet"
        _write_parquet(p)
        spec = DataSourceSpec(source_type=DataSourceType.PARQUET, path=p, options={})
        handle = ParquetConnector().open(spec)
        try:
            with pytest.raises(ValueError, match="n must be >= 1"):
                handle.read_sample(0)
        finally:
            handle.close()

    def test_sampled_fingerprint(self, tmp_path: Path) -> None:
        p = tmp_path / "t.parquet"
        _write_parquet(p)
        spec = DataSourceSpec(source_type=DataSourceType.PARQUET, path=p, options={})
        handle = ParquetConnector().open(spec)
        try:
            fp = handle.fingerprint(mode="sampled")
            assert fp.fingerprint_type == "sampled"
            assert fp.file_sha256 is None
            assert fp.content_sample_hash is not None
        finally:
            handle.close()

    def test_warnings_cached_and_capped(self, tmp_path: Path) -> None:
        p = tmp_path / "many.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["note"])
        for i in range(120):
            ws.append([f"=bad{i}"])
        wb.save(p)
        spec = DataSourceSpec(source_type=DataSourceType.XLSX, path=p, options={})
        handle = XlsxConnector().open(spec)
        try:
            first = handle.warnings()
            second = handle.warnings()
            assert len(first) == len(second) == 100
            assert first == second
            assert all(w.row >= 0 for w in first)
        finally:
            handle.close()

    def test_use_after_close_raises(self, tmp_path: Path) -> None:
        p = tmp_path / "t.parquet"
        _write_parquet(p)
        spec = DataSourceSpec(source_type=DataSourceType.PARQUET, path=p, options={})
        handle = ParquetConnector().open(spec)
        handle.close()
        with pytest.raises(ValueError, match="handle is closed"):
            handle.count_rows()

    def test_abstract_contract_not_implemented(self, tmp_path: Path) -> None:
        p = tmp_path / "t.parquet"
        _write_parquet(p)
        spec = DataSourceSpec(source_type=DataSourceType.PARQUET, path=p, options={})
        handle = FileDataHandle(spec)
        try:
            assert handle.source_type == DataSourceType.PARQUET
            assert handle.source_path == p
            with pytest.raises(NotImplementedError):
                handle._ensure_view()
            with pytest.raises(NotImplementedError):
                next(handle.read_batches())
        finally:
            handle.close()


class TestRegistry:
    def test_default_registry_has_file_connectors(self) -> None:
        registry = default_registry()
        ids = {c.connector_id for c in registry.list()}
        assert {"csv", "parquet", "jsonl", "xlsx"} <= ids

    def test_get_for_dispatch(self, tmp_path: Path) -> None:
        registry = default_registry()
        for source_type in (
            DataSourceType.CSV,
            DataSourceType.PARQUET,
            DataSourceType.JSONL,
            DataSourceType.XLSX,
        ):
            spec = DataSourceSpec(source_type=source_type, path=tmp_path / "x", options={})
            connector = registry.get_for(spec)
            assert connector.connector_id == source_type.value
        unsupported = DataSourceSpec(source_type=DataSourceType.SQLITE, path=tmp_path / "x.db")
        try:
            registry.get_for(unsupported)
            raise AssertionError("expected UnsupportedFormatError")
        except UnsupportedFormatError:
            pass

    def test_csv_still_registered(self) -> None:
        assert CsvConnector().connector_id == "csv"
